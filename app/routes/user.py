import openpyxl
from flask import Blueprint, request

from app.decorators.auth import auth_required
from app.models.user import User
from app.utils.cypher import encode_password

user = Blueprint('user', __name__, url_prefix='/user')


@user.route('/all', methods=['GET'])
@auth_required('hr')
def get_all_users():
    try:
        return {'users': [{'code': usr.code, 'name': usr.full_name, 'sector': usr.sector} for usr in
                          User.get_all()]}, 200
    except Exception as e:
        return {'message': str(e)}, 500


@user.route('/batch_register', methods=['POST'])
@auth_required('admin')
def create_many_users():
    try:
        file = request.files.get('sheet')
        if not file:
            return {"message": "Documento de planilha ausente."}, 400
        try:
            workbook = openpyxl.load_workbook(file)
        except:
            return {"message": "Planilha inválida ou corrompida."}, 400
        sheet = workbook.active
        response_draft = []
        created_quantity = 0
        error_quantity = 0
        for row in sheet.iter_rows(0, sheet.max_row):
            try:
                if not row[0].value:
                    continue
                if User.get_user(str(row[0].value)):
                    response_draft.append({row[0].value: "Usuário já existente."})
                    error_quantity += 1
                    continue
                usr = User(code=str(row[0].value),
                           full_name=str(row[1].value),
                           password=encode_password((str(row[2].value))),
                           birth_date=int(row[3].value),
                           city=str(row[4].value),
                           uf=str(row[5].value),
                           sector=str(row[6].value),
                           permission=int(row[7].value))
                usr.save()
                response_draft.append({row[0].value: "Usuário criado."})
                created_quantity += 1
            except Exception as e:
                response_draft.append(
                    {row[0].value: str(e)})
                error_quantity += 1

        return {'created_quantity': created_quantity,
                'error_quantity': error_quantity,
                'creation_info': response_draft}, 200
    except Exception as e:
        return {'message': str(e)}, 500
